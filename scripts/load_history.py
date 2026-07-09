# -*- coding: utf-8 -*-
"""
엑셀(260630 기준)의 과거 기록을 amos_posts / amos_daily_exposure에 적재한다.

  python load_history.py dry     # 쓰지 않고 계획만 출력 (기본)
  python load_history.py apply   # 실제 적재

원칙
- DELETE 절대 사용 안 함. INSERT / PATCH / upsert(ignoreDuplicates)만 사용.
- 기존 26행의 status/progress/blog_url 등 기존 값은 건드리지 않는다.
  past_urls / past_image_host_urls / past_hwaseon_urls / views_base 만 채운다.
- 신규 행은 status='미노출', progress='작업완료' (네이버 실조회로 전부 미노출 확인됨).
- views_base = 라이브로 못 구하는 조회수만. 라이브로 구할 수 있으면 넣지 않는다(이중계상 방지).
"""
import os, sys, io, json, re, urllib.request
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor

BASE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, BASE)
import backup_restore as br
sys.stdout.reconfigure(encoding='utf-8')

XLSX = os.path.join(os.path.expanduser('~'), 'Desktop',
                    '260706-아모스아윤채_상위노출 및 발행 자료(260630기준).xlsx')

DESKTOP_UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
MOBILE_UA  = 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.0 Mobile/15E148 Safari/604.1'

# 시트 표기 → DB 키워드 (사용자 확인된 별칭만)
ALIAS = {
    ('아윤채', '샴푸추천(월보장)'): '샴푸추천',
    ('아윤채', '탈모샴푸(월보장)'): '탈모샴푸',
    ('아윤채', '헤어에센스 추천(월보장)'): '헤어에센스 추천',
    ('아모스', '머릿결좋아지는법'): '머리결좋아지는법',   # 사용자 확인: 같은 키워드
}
EXCLUDE = {('아모스', '어반트라이브')}                     # 제품명 없음 → 사용자 지시로 제외

# 삭제된 카페글이라 라이브가 죽었는데 DB에 저장값이 남은 행 → base로 보존 (사용자 승인)
DEAD_CAFE_KEEP = {('아윤채', '케라스타즈 시몽'): 131,
                  ('아윤채', '탈모샴푸'): 44,
                  ('아모스', '두피가려움샴푸'): 31}


def log(m): print(f'[{__import__("datetime").datetime.now():%Y-%m-%d %H:%M:%S}] {m}')
def sp(s): return ' '.join(str(s or '').split())
def nk(s): return ''.join(str(s or '').split()).lower()
def nu(u): return str(u or '').strip().replace('https://m.blog', 'https://blog').replace('https://m.cafe', 'https://cafe').rstrip('/')
def ck(b, k): return ALIAS.get((sp(b), sp(k)), sp(k))
def is_cafe(u): return 'cafe.naver' in str(u or '')


# ---------- 라이브 조회 (lib/views.ts 와 동일 절차) ----------
def _get(url, ua, accept='*/*', referer=None):
    h = {'User-Agent': ua, 'Accept': accept}
    if referer: h['Referer'] = referer
    with urllib.request.urlopen(urllib.request.Request(url, headers=h), timeout=20) as f:
        return f.read()

def cafe_read_count(url):
    m = re.search(r'cafe\.naver\.com/([^/?#]+)/(\d+)', nu(url))
    if not m: return None
    try:
        body = _get(f'https://cafe.naver.com/{m.group(1)}/{m.group(2)}', DESKTOP_UA, referer='https://cafe.naver.com/')
        cm = re.search(rb'clubid=(\d+)', body)
        if not cm: return None
        j = json.loads(_get(
            f'https://apis.naver.com/cafe-web/cafe-articleapi/v2.1/cafes/{cm.group(1).decode()}/articles/{m.group(2)}?query=&menuId=0&boardType=L',
            MOBILE_UA, 'application/json', 'https://m.cafe.naver.com/'))
        rc = j.get('result', {}).get('article', {}).get('readCount')
        return rc if isinstance(rc, int) else None
    except Exception as e:
        print(f'  오류: 카페 조회 실패 {url} - {e}')
        return None

def image_views(u):
    m = re.search(r'hwaseon-image\.com/(?:image|uploads)/([a-zA-Z0-9_-]+)', str(u or ''))
    if not m: return None
    try:
        j = json.loads(_get(f'https://hwaseon-image.com/image/{m.group(1)}/detail', DESKTOP_UA, 'application/json'))
        v = j.get('views')
        return v if isinstance(v, int) else None
    except Exception as e:
        print(f'  오류: 이미지 조회 실패 {u} - {e}')
        return None


# ---------- 엑셀 파싱 ----------
def load_excel():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    COLS = ['구분','브랜드','제품','키워드','검색량','노출탭','진행상태','비고','시작일','노출일',
            '발행URL','발행일','블로거','이미지호스팅링크','조회수','제품URL','클릭수']
    def num(v):
        if v in (None, '', '-'): return None
        try: return int(float(str(v).replace(',', '')))
        except Exception: return None
    rows = []
    for name, skip in [('~260630발행마스터', 1), ('Sheet1', 0)]:
        ws = wb[name]
        for r in ws.iter_rows(min_row=skip+1, values_only=True):
            if not any(r): continue
            d = dict(zip(COLS, [c if not isinstance(c, str) else c.strip() for c in r]))
            d['조회수'] = num(d['조회수']); d['클릭수'] = num(d['클릭수'])
            rows.append(d)

    # 일자별 노출현황
    ws = wb['~260630일자별노출현황']
    hdr = list(ws.iter_rows(values_only=True))[0]
    import datetime as dt
    dates, year, prev_m = [], 2025, 0
    for i, h in enumerate(hdr[6:], start=6):
        m, dd = [int(x) for x in str(h).split('/')]
        if prev_m and m < prev_m: year += 1
        prev_m = m
        dates.append((i, dt.date(year, m, dd)))
    expo = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not any(r): continue
        days = [str(d) for i, d in dates if r[i] not in (None, '')]
        expo.append({'brand': sp(r[0]), 'keyword': ck(r[0], r[1]), 'days': days})
    return rows, expo


def main(apply_mode):
    url, key = br.load_env()
    log(f"모드: {'APPLY (실제 적재)' if apply_mode else 'DRY-RUN (쓰기 없음)'}")

    if apply_mode:
        log('적재 전 백업 생성...')
        br.do_backup()

    ex, expo = load_excel()
    db = br.fetch_all(url, key, 'amos_posts')
    dbrow = {(nk(r['brand']), nk(r['keyword'])): r for r in db}
    log(f'DB {len(db)}행 / 엑셀 {len(ex)}행 / 노출시트 {len(expo)}행')

    # 키워드별 집계
    G = defaultdict(lambda: {'ex': [], 'days': set()})
    for r in ex:
        b, kw = sp(r['브랜드']), ck(r['브랜드'], r['키워드'])
        if (b, kw) in EXCLUDE: continue
        g = G[(nk(b), nk(kw))]; g.setdefault('disp', (b, kw)); g['ex'].append(r)
    for r in expo:
        b, kw = r['brand'], r['keyword']
        if (b, kw) in EXCLUDE: continue
        g = G[(nk(b), nk(kw))]; g.setdefault('disp', (b, kw)); g['days'] |= set(r['days'])

    # 각 그룹의 URL을 발행일 최신순 정렬 (무일자는 가장 오래된 것으로)
    for k, g in G.items():
        seen, urls = set(), []
        for r in g['ex']:
            u = nu(r['발행URL'])
            if not u or u in seen: continue
            seen.add(u); urls.append(r)
        urls.sort(key=lambda r: sp(r['발행일'])[:10] or '0000-00-00', reverse=True)
        g['urls'] = urls
        g['product'] = next((sp(r['제품']) for r in g['ex'] if r['제품']), '')
        g['tab'] = next((sp(r['노출탭']) for r in g['ex'] if r['노출탭']), '')

    # 라이브 조회 (중복 제거)
    cafes = sorted({nu(r['발행URL']) for g in G.values() for r in g['urls'] if is_cafe(r['발행URL'])}
                   | {nu(r['blog_url']) for r in db if is_cafe(r['blog_url'])})
    imgs = sorted({sp(r['이미지호스팅링크']) for g in G.values() for r in g['ex'] if r['이미지호스팅링크']}
                  | {sp(r['image_host_url']) for r in db if r['image_host_url']})
    log(f'라이브 조회: 카페 {len(cafes)}건 / 이미지 {len(imgs)}건')
    with ThreadPoolExecutor(max_workers=8) as p:
        cafe_v = dict(zip(cafes, p.map(cafe_read_count, cafes)))
        img_v = dict(zip(imgs, p.map(image_views, imgs)))

    plans = []
    for k, g in G.items():
        b, kw = g['disp']
        d = dbrow.get(k)

        # --- URL 배분 ---
        excel_urls = [nu(r['발행URL']) for r in g['urls']]
        if d:
            cur_url = nu(d['blog_url']) if d['blog_url'] else (excel_urls[0] if excel_urls else None)
            past = [u for u in excel_urls if u != cur_url]              # 기존 blog_url 유지
        else:
            cur_url = excel_urls[0] if excel_urls else None             # 최신 = 첫 번째
            past = excel_urls[1:]

        # --- 이미지호스팅 / 단축코드: 현재 URL 것을 대표로, 나머지는 past ---
        # ⚠ 발행URL 없이 이미지/단축코드만 있는 행(비듬샴푸 등)이 있으므로 g['ex'] 전체에서 수집한다.
        #    URL 있는 행을 먼저 훑어 순서를 유지하고, 그 뒤 나머지 행을 붙인다.
        img_by_url = {nu(r['발행URL']): sp(r['이미지호스팅링크']) for r in g['urls'] if r['이미지호스팅링크']}
        sh_by_url = {nu(r['발행URL']): sp(r['제품URL']) for r in g['urls'] if r['제품URL']}
        all_imgs = list(dict.fromkeys([sp(r['이미지호스팅링크']) for r in g['urls'] if r['이미지호스팅링크']]
                                      + [sp(r['이미지호스팅링크']) for r in g['ex'] if r['이미지호스팅링크']]))
        all_shs = list(dict.fromkeys([sp(r['제품URL']) for r in g['urls'] if r['제품URL']]
                                     + [sp(r['제품URL']) for r in g['ex'] if r['제품URL']]))
        cur_img = (sp(d['image_host_url']) if d and d['image_host_url']
                   else img_by_url.get(cur_url) or (all_imgs[0] if all_imgs else ''))
        cur_sh = (sp(d['hwaseon_url']) if d and d['hwaseon_url']
                  else sh_by_url.get(cur_url) or (all_shs[0] if all_shs else ''))
        past_imgs = [v for v in all_imgs if v != cur_img]
        past_shs = [v for v in all_shs if v != cur_sh]

        # --- views_base: 라이브로 못 구하는 조회수만 ---
        base, why, seen_img = 0, [], set()
        for r in g['ex']:
            v = r['조회수']
            if v is None: continue
            u = nu(r['발행URL']); img = sp(r['이미지호스팅링크'])
            if is_cafe(u):
                live = cafe_v.get(u)
            elif img:
                if img in seen_img: continue          # 같은 이미지 중복 합산 방지
                seen_img.add(img)
                live = img_v.get(img)
            else:
                live = None
            if live is None:
                base += v; why.append(f'{v}(라이브불가)')
            elif live < v:
                base += v; why.append(f'{v}(소스리셋 live={live})')
        # 삭제된 카페글의 DB 저장값 보존. 셋 다 엑셀의 같은 URL을 가리키므로 더하지 않고 max를 취한다.
        keep = DEAD_CAFE_KEEP.get((b, kw))
        if keep is not None:
            base = max(base, keep)
            why.append(f'{keep}(삭제카페글 저장값 보존)')

        # 기존 행에서 비어 있던 칸만 새로 채운다(값이 있으면 절대 덮어쓰지 않음).
        # 이 처리를 안 하면 cur_url/cur_img/cur_sh 가 past에서 제외된 채 어디에도 저장되지 않아 유실된다.
        fill = {}
        if d:
            if not d['blog_url'] and cur_url: fill['blog_url'] = cur_url
            if not d['image_host_url'] and cur_img: fill['image_host_url'] = cur_img
            if not d['hwaseon_url'] and cur_sh: fill['hwaseon_url'] = cur_sh

        plans.append({'key': k, 'brand': b, 'keyword': kw, 'is_new': d is None, 'id': d['id'] if d else None,
                      'product': (d['product'] if d else g['product']), 'tab': g['tab'],
                      'blog_url': cur_url, 'past_urls': past,
                      'image_host_url': cur_img, 'past_image_host_urls': past_imgs,
                      'hwaseon_url': cur_sh, 'past_hwaseon_urls': past_shs,
                      'views_base': base, 'why': why, 'days': sorted(g['days']), 'fill': fill})

    # ---------- 출력 ----------
    new = [p for p in plans if p['is_new']]
    old = [p for p in plans if not p['is_new']]
    print(f'\n=== 신규 생성 {len(new)}행 (status=미노출, progress=작업완료) ===')
    for p in sorted(new, key=lambda x: -x['views_base']):
        print(f"  {p['brand']} {p['keyword'][:16]:18} 제품={p['product'][:22]:24} base={p['views_base']:>5} "
              f"URL={'있음' if p['blog_url'] else '없음'} past={len(p['past_urls'])} img={len(p['past_image_host_urls'])} sh={len(p['past_hwaseon_urls'])} 노출일={len(p['days'])}")
    noprod = [p for p in new if not p['product']]
    if noprod:
        print(f"  ⚠ 제품명 없음(생성 불가): {[(p['brand'], p['keyword']) for p in noprod]}")

    print(f'\n=== 기존 {len(old)}행 갱신 (past_* + views_base + 빈칸 채움) ===')
    for p in sorted(old, key=lambda x: -x['views_base']):
        if p['past_urls'] or p['past_image_host_urls'] or p['past_hwaseon_urls'] or p['views_base'] or p['fill']:
            fl = ('  빈칸채움=' + ','.join(p['fill'])) if p['fill'] else ''
            print(f"  {p['brand']} {p['keyword'][:16]:18} base={p['views_base']:>5} past_url={len(p['past_urls'])} "
                  f"past_img={len(p['past_image_host_urls'])} past_sh={len(p['past_hwaseon_urls'])} 노출일={len(p['days'])}{fl}")

    # 유실 검증: 엑셀의 모든 URL/이미지/단축코드가 현재값 또는 past_* 중 한 곳에 반드시 남아야 한다
    print('\n=== 유실 검증 ===')
    lost = 0
    for p in plans:
        g = G[p['key']]
        d = dbrow.get(p['key'])
        cur = p['fill'].get('blog_url') or (nu(d['blog_url']) if d and d['blog_url'] else p['blog_url'])
        stored = set(p['past_urls']) | ({cur} if cur else set())
        miss = {nu(r['발행URL']) for r in g['urls']} - stored
        if miss:
            lost += len(miss); print(f"  URL 유실! {p['brand']} {p['keyword']}: {miss}")

        for label, curv, pastv, srcs in [
            ('이미지', p['image_host_url'], p['past_image_host_urls'], {sp(r['이미지호스팅링크']) for r in g['ex'] if r['이미지호스팅링크']}),
            ('단축코드', p['hwaseon_url'], p['past_hwaseon_urls'], {sp(r['제품URL']) for r in g['ex'] if r['제품URL']})]:
            st = set(pastv) | ({curv} if curv else set())
            m2 = srcs - st
            if m2:
                lost += len(m2); print(f"  {label} 유실! {p['brand']} {p['keyword']}: {m2}")
    print(f'  유실 총 {lost}건' + ('' if lost else ' ✅'))

    tot_days = sum(len(p['days']) for p in plans)
    print(f"\nviews_base 합계 {sum(p['views_base'] for p in plans):,} / 노출일 총 {tot_days}건")

    if not apply_mode:
        print('\nDRY-RUN 종료 — DB 변경 없음')
        return

    # ---------- 실제 적재 ----------
    if noprod:
        print('중단: 제품명 없는 신규행이 있습니다.')
        sys.exit(1)

    log('신규 행 생성...')
    for p in new:
        body = {'brand': p['brand'], 'product': p['product'], 'keyword': p['keyword'],
                'status': '미노출', 'progress': '작업완료',
                'tab_type': p['tab'] or None, 'blog_url': p['blog_url'],
                'past_urls': ', '.join(p['past_urls']) or None,
                'image_host_url': p['image_host_url'] or None,
                'past_image_host_urls': ', '.join(p['past_image_host_urls']) or None,
                'hwaseon_url': p['hwaseon_url'] or None,
                'past_hwaseon_urls': ', '.join(p['past_hwaseon_urls']) or None,
                'views_base': p['views_base'], 'views_offset': 0}
        res = br.req(url, key, 'amos_posts', method='POST', body=body, prefer='return=representation')
        p['id'] = res[0]['id']
    log(f'신규 {len(new)}행 생성 완료')

    log('기존 행 갱신 (past_* + views_base)...')
    for p in old:
        body = {'past_urls': ', '.join(p['past_urls']) or None,
                'past_image_host_urls': ', '.join(p['past_image_host_urls']) or None,
                'past_hwaseon_urls': ', '.join(p['past_hwaseon_urls']) or None,
                'views_base': p['views_base'], 'views_offset': 0}
        body.update(p['fill'])   # 비어 있던 blog_url/image_host_url/hwaseon_url 만 채움
        br.req(url, key, f"amos_posts?id=eq.{p['id']}", method='PATCH', body=body)
    log(f'기존 {len(old)}행 갱신 완료')

    log('노출일 upsert...')
    recs = [{'post_id': p['id'], 'date': d} for p in plans for d in p['days']]
    for i in range(0, len(recs), 500):
        br.req(url, key, 'amos_daily_exposure', method='POST', body=recs[i:i+500],
               prefer='resolution=ignore-duplicates')
    log(f'노출일 {len(recs)}건 upsert 완료')

    exp = br.fetch_all(url, key, 'amos_daily_exposure')
    posts = br.fetch_all(url, key, 'amos_posts')
    log(f'최종: amos_posts {len(posts)}행 / amos_daily_exposure {len(exp)}건')


if __name__ == '__main__':
    main(len(sys.argv) > 1 and sys.argv[1] == 'apply')
